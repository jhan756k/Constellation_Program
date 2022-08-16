from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from openpyxl import load_workbook
import random, time

load_wb = load_workbook("Constellation.xlsx", data_only=True)
constellations = load_wb['시트1']
qtype = ("한국어명 Constellation Abbreviation Alpha Beta Gamma").split()

spring_const = [[0]*(6) for _ in range(12)]
summer_const = [[0]*(6) for _ in range(13)]
autumn_const = [[0]*(6) for _ in range(12)]
winter_const = [[0]*(6) for _ in range(10)]
north_const = [[0]*(6) for _ in range(6)]
south_const = [[0]*(6) for _ in range(34)]
reslist = []
const_len = [12, 13, 12, 10, 6, 34]
total_const = [spring_const, summer_const, autumn_const, winter_const, north_const, south_const]

for springs in range(12):
    for x in range(6):
        spring_const[springs][x] = constellations.cell(row=springs+4, column=x+1).value

for summers in range(13):
    for x in range(6):
        summer_const[summers][x] = constellations.cell(row=summers+21, column=x+1).value

for autumn in range(12):
    for x in range(6):
        autumn_const[autumn][x] = constellations.cell(row=autumn+39, column=x+1).value

for winter in range(10):
    for x in range(6):
        winter_const[winter][x] = constellations.cell(row=winter+57, column=x+1).value

for north in range(6):
    for x in range(6):
        north_const[north][x] = constellations.cell(row=north+72, column=x+1).value

for south in range(34):
    for x in range(6):
        south_const[south][x] = constellations.cell(row=south+83, column=x+1).value

uselist = []
rmax = 0
qtext = ""
jechul = False
jechulgap = ""
daum = False

class Thread1(QThread):
    qtxt_signal = pyqtSignal(str)
    anstxt_signal = pyqtSignal(str)
    resetinput_sig = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)

    def run(self):
        global rmax, qtext, jechul, daum
        
        for x in range(6):
            if x in reslist:
                uselist.extend(total_const[x])
                rmax += const_len[x]

        while True:

            r = random.randint(0, rmax-1)
            anstype = random.randint(1, 5)

            if (uselist[r][anstype] == None) or (uselist[r][anstype] == " ") or (uselist[r][anstype] == "-"):
                continue
            
            qtext = f"{uselist[r][0]} 의 {qtype[anstype]} 은 무엇일까요?"
            self.qtxt_signal.emit(qtext)

            while not jechul:
                time.sleep(1)

            jechul = False

            if jechulgap == uselist[r][anstype]:
                self.anstxt_signal.emit("맞습니다.")
                time.sleep(1)

            else:
                self.anstxt_signal.emit(f"틀렸습니다. 정답은 {uselist[r][anstype]} 입니다.")
                
                while not daum:
                    time.sleep(1)

                daum = False

            self.anstxt_signal.emit("")
            self.resetinput_sig.emit()

class Ui_MainWindow(object):

    def setqtxt(self, txt):
        self.QuestionLabel.setText(txt)

    def setatxt(self, txt):
        self.CheckansLabel.setText(txt)

    def checkdaum(self):
        global daum
        daum = True
    
    def resetinp(self):
        self.Ansinput.setText("")

    def checkans(self):
        global jechul, jechulgap
        jechul = True
        jechulgap = self.Ansinput.text()

    def start(self):
        checkboxlist = [self.Springcheck, self.Summercheck, self.Autumncheck, self.Wintercheck, self.Northcheck, self.Southcheck]

        if all(not checkboxlist[a].isChecked() for a in range(6)):
            return

        for x in range(6):
            if checkboxlist[x].isChecked():
                reslist.append(x)

        self.startButton.setEnabled(False)

        for x in range(6):
            checkboxlist[x].setEnabled(False)

        x = Thread1(MainWindow)
        x.start()
        x.qtxt_signal.connect(self.setqtxt)
        x.anstxt_signal.connect(self.setatxt)
        x.resetinput_sig.connect(self.resetinp)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(819, 485)
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(20)
        font.setBold(False)
        font.setWeight(50)
        MainWindow.setFont(font)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.gridLayoutWidget.setGeometry(QtCore.QRect(10, 10, 121, 217))
        self.gridLayoutWidget.setObjectName("gridLayoutWidget")
        self.gridLayout = QtWidgets.QGridLayout(self.gridLayoutWidget)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")
        self.Northcheck = QtWidgets.QCheckBox(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.Northcheck.setFont(font)
        self.Northcheck.setObjectName("Northcheck")
        self.gridLayout.addWidget(self.Northcheck, 4, 0, 1, 1)
        self.Autumncheck = QtWidgets.QCheckBox(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.Autumncheck.setFont(font)
        self.Autumncheck.setObjectName("Autumncheck")
        self.gridLayout.addWidget(self.Autumncheck, 2, 0, 1, 1)
        self.Springcheck = QtWidgets.QCheckBox(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.Springcheck.setFont(font)
        self.Springcheck.setObjectName("Springcheck")
        self.gridLayout.addWidget(self.Springcheck, 0, 0, 1, 1)
        self.Southcheck = QtWidgets.QCheckBox(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.Southcheck.setFont(font)
        self.Southcheck.setObjectName("Southcheck")
        self.gridLayout.addWidget(self.Southcheck, 5, 0, 1, 1)
        self.Wintercheck = QtWidgets.QCheckBox(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.Wintercheck.setFont(font)
        self.Wintercheck.setObjectName("Wintercheck")
        self.gridLayout.addWidget(self.Wintercheck, 3, 0, 1, 1)
        self.Summercheck = QtWidgets.QCheckBox(self.gridLayoutWidget)
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.Summercheck.setFont(font)
        self.Summercheck.setObjectName("Summercheck")
        self.gridLayout.addWidget(self.Summercheck, 1, 0, 1, 1)
        self.startButton = QtWidgets.QPushButton(self.centralwidget)
        self.startButton.setGeometry(QtCore.QRect(10, 240, 111, 41))
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.startButton.setFont(font)
        self.startButton.setObjectName("startButton")
        self.QuestionLabel = QtWidgets.QLabel(self.centralwidget)
        self.QuestionLabel.setGeometry(QtCore.QRect(180, 10, 621, 91))
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.QuestionLabel.setFont(font)
        self.QuestionLabel.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.QuestionLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.QuestionLabel.setWordWrap(True)
        self.QuestionLabel.setIndent(4)
        self.QuestionLabel.setObjectName("QuestionLabel")
        self.Ansinput = QtWidgets.QLineEdit(self.centralwidget)
        self.Ansinput.setGeometry(QtCore.QRect(180, 170, 351, 61))
        self.Ansinput.setObjectName("Ansinput")
        self.CheckansLabel = QtWidgets.QLabel(self.centralwidget)
        self.CheckansLabel.setGeometry(QtCore.QRect(180, 260, 411, 111))
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.CheckansLabel.setFont(font)
        self.CheckansLabel.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignVCenter)
        self.CheckansLabel.setWordWrap(True)
        self.CheckansLabel.setObjectName("CheckansLabel")
        self.enteransbutton = QtWidgets.QPushButton(self.centralwidget)
        self.enteransbutton.setGeometry(QtCore.QRect(580, 170, 121, 61))
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(30)
        font.setBold(True)
        font.setWeight(75)
        self.enteransbutton.setFont(font)
        self.enteransbutton.setObjectName("enteransbutton")
        self.nextbutton = QtWidgets.QPushButton(self.centralwidget)
        self.nextbutton.setGeometry(QtCore.QRect(650, 280, 121, 61))
        font = QtGui.QFont()
        font.setFamily("NanumGothic")
        font.setPointSize(30)
        font.setBold(True)
        font.setWeight(75)
        self.nextbutton.setFont(font)
        self.nextbutton.setObjectName("nextbutton")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 819, 44))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        MainWindow.setWindowTitle("Constellation Program - made by 한준희")
        self.Northcheck.setText("North")
        self.Autumncheck.setText("Autumn")
        self.Springcheck.setText("Spring")
        self.Southcheck.setText("South")
        self.Wintercheck.setText("Winter")
        self.Summercheck.setText("Summer")
        self.startButton.setText("START")
        self.QuestionLabel.setText("Question")
        self.CheckansLabel.setText("Answer")
        self.enteransbutton.setText("제출")
        self.nextbutton.setText("다음")
        
        self.startButton.clicked.connect(self.start)
        self.enteransbutton.clicked.connect(self.checkans)
        self.Ansinput.returnPressed.connect(self.checkans)
        self.nextbutton.clicked.connect(self.checkdaum)

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
