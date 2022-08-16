from openpyxl import load_workbook
import random
load_wb = load_workbook("Constellation.xlsx", data_only=True)
constellations = load_wb['시트1']
qtype = ("한국어명 Constellation Abbreviation Alpha Beta Gamma").split()

spring_const = [[0]*(6) for _ in range(12)]
summer_const = [[0]*(6) for _ in range(13)]
autumn_const = [[0]*(6) for _ in range(12)]
winter_const = [[0]*(6) for _ in range(10)]
north_const = [[0]*(6) for _ in range(6)]
south_const = [[0]*(6) for _ in range(34)]

const_len = [12, 13, 12, 10, 6, 34]
total_const = [spring_const, summer_const, autumn_const, winter_const, north_const, south_const]

for springs in range(12):
    for x in range(6):
        spring_const[springs][x] = constellations.cell(row=springs+4, column=x+1).value

for summers in range(13):
    for x in range(6):
        summer_const[summers][x] = constellations.cell(row=summers+21, column=x+1).value

for autumn in range(12):
    for x in range(6):
        autumn_const[autumn][x] = constellations.cell(row=autumn+39, column=x+1).value

for winter in range(10):
    for x in range(6):
        winter_const[winter][x] = constellations.cell(row=winter+57, column=x+1).value

for north in range(6):
    for x in range(6):
        north_const[north][x] = constellations.cell(row=north+72, column=x+1).value

for south in range(34):
    for x in range(6):
        south_const[south][x] = constellations.cell(row=south+83, column=x+1).value

print("spring 1, summer 2, autumn 3, winter 4, north 5, south 6 중 고르세요")

getuse = list(map(int, input().split()))
uselist = []
rmax = 0
for x in range(1, 7):
    if x in getuse:
        uselist.extend(total_const[x-1])
        rmax += const_len[x-1]
print()
print("멈추고 싶으면 stop 이라고 치세요")
print()

while True:
    r = random.randint(0, rmax-1)
    anstype = random.randint(1, 5)

    if (uselist[r][anstype] == None) or (uselist[r][anstype] == " ") or (uselist[r][anstype] == "-"):
        continue

    print(f"{uselist[r][0]} 의 {qtype[anstype]} 은 무엇일까요?")
    getans = str(input())

    if getans == "stop":
        exit()
    
    if getans == uselist[r][anstype]:
        print("맞습니다.", end = "\n\n")

    else:
        print("틀렸습니다.")
        print(f"정답은 {uselist[r][anstype]} 입니다.", end = "\n\n")
